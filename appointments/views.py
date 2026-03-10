import openpyxl
from django.shortcuts import render, redirect
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from rest_framework import viewsets

# Importación de la función de utilidad para WhatsApp
from .utils import enviar_confirmacion_whatsapp

# Importaciones de modelos y formularios locales
from .models import Turno, Persona, Servicio, ConfiguracionSistema
from .forms import TurnoForm, PacienteForm, ServicioForm
from .serializers import PersonaSerializer, ServicioSerializer

# --- VISTA PÚBLICA (PORTADA) ---

def portada(request):
    """Página de inicio pública de Vivand."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'appointments/portada.html')

# --- VISTAS PROTEGIDAS (Requieren Login) ---

@login_required
def dashboard(request):
    """Vista principal con Ingresos Mensuales y estadísticas de Vivand."""
    ahora = timezone.now()
    
    total_ingresos_mes = Turno.objects.filter(
        estado='COMPLETADO',
        inicio__month=ahora.month,
        inicio__year=ahora.year
    ).aggregate(total=Sum('servicio__precio'))['total'] or 0
    
    servicios_mes = Turno.objects.filter(
        estado='COMPLETADO',
        inicio__month=ahora.month,
        inicio__year=ahora.year
    ).count()
    
    proximos_turnos = Turno.objects.filter(
        inicio__gte=ahora
    ).order_by('inicio')[:5]
    
    servicios_populares = Servicio.objects.annotate(
        num_turnos=Count('turno')
    ).order_by('-num_turnos')[:5]

    return render(request, 'appointments/dashboard.html', {
        'total_ingresos': total_ingresos_mes,
        'total_servicios': servicios_mes,
        'proximos_turnos': proximos_turnos,
        'servicios_populares': servicios_populares,
        'mes_actual': ahora.strftime('%B'),
    })

@login_required
def nuevo_turno(request):
    """Vista para agendar un nuevo turno con notificación automática vía UltraMsg."""
    if request.method == 'POST':
        form = TurnoForm(request.POST)
        if form.is_valid():
            # 1. Guardamos el objeto en la base de datos
            turno_guardado = form.save()
            
            # 2. Obtenemos los datos del paciente relacionado
            paciente = turno_guardado.paciente
            
            # 3. Validamos que el paciente tenga teléfono antes de enviar
            if paciente and hasattr(paciente, 'telefono') and paciente.telefono:
                nombre_completo = f"{paciente.nombre} {paciente.apellido}"
                # Formateamos la fecha para que sea legible en el mensaje
                fecha_hora = turno_guardado.inicio.strftime('%d/%m/%Y a las %H:%M')
                
                enviar_confirmacion_whatsapp(
                    telefono=paciente.telefono, 
                    nombre_paciente=nombre_completo,
                    hora_turno=fecha_hora
                )
                messages.success(request, "Turno agendado y notificación enviada por WhatsApp.")
            else:
                messages.warning(request, "Turno agendado, pero el paciente no tiene teléfono registrado.")

            return redirect('agenda')
    else:
        form = TurnoForm()
    
    return render(request, 'appointments/formulario_generico.html', {
        'form': form,
        'titulo': 'Agendar Nuevo Turno'
    })

@login_required
def crear_persona(request):
    """Vista para crear una nueva Persona (Paciente/Cuidador)."""
    if request.method == 'POST':
        form = PacienteForm(request.POST)
        if form.is_valid():
            persona = form.save(commit=False)
            if not persona.tipo:
                persona.tipo = 'PACIENTE'
            persona.save()
            messages.success(request, "Persona registrada exitosamente.")
            return redirect('dashboard')
    else:
        form = PacienteForm()
    return render(request, 'appointments/formulario_generico.html', {
        'form': form, 
        'titulo': 'Registrar Nueva Persona'
    })

@login_required
def crear_servicio(request):
    """Vista para crear un nuevo Servicio."""
    if request.method == 'POST':
        form = ServicioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Servicio configurado correctamente.")
            return redirect('dashboard')
    else:
        form = ServicioForm()
    return render(request, 'appointments/formulario_generico.html', {
        'form': form, 
        'titulo': 'Configurar Nuevo Servicio'
    })

@login_required
def exportar_turnos_excel(request):
    """Genera un Informe de Ficha Médica en Excel."""
    query_paciente = request.GET.get('q_paciente') or request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    turnos = Turno.objects.all().order_by('-inicio')

    if query_paciente:
        turnos = turnos.filter(
            Q(paciente__nombre__icontains=query_paciente) | 
            Q(paciente__apellido__icontains=query_paciente)
        )
    
    if fecha_inicio and fecha_fin:
        turnos = turnos.filter(inicio__date__range=[fecha_inicio, fecha_fin])

    if not turnos.exists():
        messages.warning(request, "No se encontraron datos para generar la ficha clínica.")
        return redirect('dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ficha Médica"

    fuente_titulo = openpyxl.styles.Font(bold=True, size=14, color="198754")
    fuente_subtitulo = openpyxl.styles.Font(bold=True, size=11)
    relleno_gris = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.merge_cells('A1:E1')
    ws['A1'] = "VIVAND - GESTIÓN DE CUIDADOS INTEGRALES"
    ws['A1'].font = fuente_titulo
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

    p = turnos.first().paciente
    ws.append([])
    ws.append(['I. ANTECEDENTES DEL PACIENTE'])
    ws[f'A{ws.max_row}'].font = fuente_subtitulo
    ws[f'A{ws.max_row}'].fill = relleno_gris

    ws.append(['NOMBRE:', f"{p.nombre} {p.apellido}".upper(), '', 'RUT:', getattr(p, 'rut', 'N/A')])
    ws.append(['DIAGNÓSTICO:', getattr(p, 'diagnostico', 'No especificado')])
    ws.append(['TRATAMIENTOS:', getattr(p, 'tratamiento_vigente', 'Sin registros')])
    ws.append([])

    ws.append(['II. REGISTRO DE ATENCIONES'])
    ws[f'A{ws.max_row}'].font = fuente_subtitulo
    ws[f'A{ws.max_row}'].fill = relleno_gris
    
    ws.append(['FECHA', 'SERVICIO', 'CUIDADOR', 'ESTADO', 'OBSERVACIONES'])
    for cell in ws[ws.max_row]:
        cell.font = fuente_subtitulo

    total_inversion = 0
    for t in turnos:
        precio = t.servicio.precio or 0
        total_inversion += precio
        ws.append([
            t.inicio.strftime('%d/%m/%Y %H:%M'),
            t.servicio.nombre,
            f"{t.cuidador.nombre} {t.cuidador.apellido}" if t.cuidador else "No asignado",
            t.get_estado_display(),
            getattr(t, 'observaciones', "")
        ])

    ws.append([])
    ws.append(['', '', '', 'TOTAL INVERSIÓN:', total_inversion])
    ws.cell(row=ws.max_row, column=5).font = fuente_subtitulo

    anchos = [20, 30, 25, 15, 40]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Ficha_{p.apellido}.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_cuidador_mensual(request):
    """Genera reporte de pagos para el cuidador."""
    ahora = timezone.now()
    mes_form = request.GET.get('mes')
    mes_actual = int(mes_form) if mes_form else ahora.month
    anio_actual = ahora.year

    turnos = Turno.objects.filter(
        inicio__month=mes_actual,
        inicio__year=anio_actual,
        estado='COMPLETADO'
    ).order_by('cuidador', 'inicio')

    if not turnos.exists():
        messages.warning(request, f"No hay servicios completados para el mes seleccionado.")
        return redirect('dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pagos_{mes_actual}_{anio_actual}"

    fuente_titulo = openpyxl.styles.Font(bold=True, size=14, color="0D6EFD")
    fuente_subtitulo = openpyxl.styles.Font(bold=True, size=11)
    relleno_azul = openpyxl.styles.PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")

    ws.merge_cells('A1:E1')
    ws['A1'] = "REPORTE DE SERVICIOS Y PAGOS - VIVAND"
    ws['A1'].font = fuente_titulo
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

    ws.append([])
    ws.append(['PERIODO:', f"{mes_actual}/{anio_actual}", '', 'FECHA EMISIÓN:', ahora.strftime('%d/%m/%Y')])
    ws.append([])
    ws.append(['CUIDADOR/A', 'FECHA SERVICIO', 'PACIENTE', 'SERVICIO REALIZADO', 'HONORARIOS'])
    
    for cell in ws[ws.max_row]:
        cell.font = fuente_subtitulo
        cell.fill = relleno_azul

    total_a_pagar = 0
    for t in turnos:
        precio = t.servicio.precio or 0
        total_a_pagar += precio
        ws.append([
            f"{t.cuidador.nombre} {t.cuidador.apellido}".upper() if t.cuidador else "SIN ASIGNAR",
            t.inicio.strftime('%d/%m/%Y'),
            f"{t.paciente.nombre} {t.paciente.apellido}",
            t.servicio.nombre,
            precio
        ])

    ws.append([])
    ws.append(['', '', '', 'TOTAL A PAGAR:', total_a_pagar])
    ws[f'E{ws.max_row}'].number_format = '"$"#,##0'

    anchos = [25, 15, 25, 35, 15]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="Pagos_Vivand_Mes_{mes_actual}.xlsx"'
    wb.save(response)
    return response

@login_required
def agenda(request):
    """Vista de la lista de turnos (Agenda)."""
    query = request.GET.get('q')
    turnos = Turno.objects.all()
    if query:
        turnos = turnos.filter(
            Q(paciente__nombre__icontains=query) | 
            Q(paciente__apellido__icontains=query) |
            Q(cuidador__nombre__icontains=query) |
            Q(titulo__icontains=query)
        )
    return render(request, 'appointments/agenda.html', {'turnos': turnos.order_by('inicio'), 'query': query})

@login_required
def calendar_events(request):
    """Eventos para el calendario FullCalendar."""
    turnos = Turno.objects.all()
    data = []
    for t in turnos:
        data.append({
            'id': t.id,
            'title': f"{t.titulo} - {t.paciente.nombre}",
            'start': t.inicio.isoformat(),
            'end': t.fin.isoformat(),
            'color': getattr(t, 'color_etiqueta', '#0d6efd'),
        })
    return JsonResponse(data, safe=False)

@login_required
def lista_personas(request):
    """Gestión de Pacientes y Cuidadores."""
    query = request.GET.get('q')
    personas = Persona.objects.all().order_by('apellido')
    
    if query:
        personas = personas.filter(
            Q(nombre__icontains=query) | 
            Q(apellido__icontains=query) |
            Q(rut__icontains=query)
        )
    
    config, _ = ConfiguracionSistema.objects.get_or_create(id=1)
    return render(request, 'appointments/lista.html', {
        'personas': personas, 
        'config': config,
        'query': query
    })

@login_required
def lista_servicios(request):
    """Lista de servicios configurados."""
    servicios = Servicio.objects.all()
    return render(request, 'appointments/lista_servicios.html', {'servicios': servicios})

def verificar_rut_disponible(request):
    """Validación AJAX de RUT."""
    rut = request.GET.get('rut', None)
    rut_limpio = rut.replace(".", "").replace("-", "").upper() if rut else ""
    existe = Persona.objects.filter(rut__icontains=rut_limpio).exists()
    return JsonResponse({'existe': existe}) 

# --- VIEWSETS PARA API ---

class PersonaViewSet(viewsets.ModelViewSet):
    queryset = Persona.objects.all()
    serializer_class = PersonaSerializer

class ServicioViewSet(viewsets.ModelViewSet):
    queryset = Servicio.objects.all()
    serializer_class = ServicioSerializer
    
@login_required
def exportar_ficha_paciente(request):
    """Función puente para reutilizar la lógica de exportación Excel."""
    return exportar_turnos_excel(request)    